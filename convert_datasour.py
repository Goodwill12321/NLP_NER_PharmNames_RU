from openpyxl import load_workbook
import json
from tqdm import tqdm

def convert_excel_to_datasaur(input_file, output_file, max_records=None):
    # Колонки, которые нужно полностью исключить из вывода
    EXCLUDED_COLUMNS = {
        "№", "МНН", "МННСтрокой", "ЖН", "СМНН", "КЛП", "Наркотический", 
        "ОКПД2", "РегНомерПредЦены", "ДатаОкончанияПредЦены", "ВидЗаписиРеестраЖН", "ДатаРегистрацииПредЦены","Дозировка_ЕИ_Потребительская_КодОКЕИ",    
        "НомерРешенияРеестраЖН", "ДатаРегистрацииЦеныРеестраЖН", "ДатаВступленияВСилуРеестраЖН", "Дозировка_ЕИ_КодОКЕИ",
        "Дозировка_ЕИ_ПотребительскаяОКЕИ", "Штрихкод", "ВладелецНаименование", 
        "ВладелецСтрана", "ЛекформаДозировкаУпаковкаИзРеестраЖН", "ВладелецРУИзРеестраЖН", 
        "НомерРУ", "ДатаРУ"
    }

    print("🔍 Чтение Excel файла...")
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb.active

    # Получаем все заголовки колонок
    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # Определяем колонки для разметки (все, кроме исключённых)
    text_columns = ["ТоварПоставки", "ПредставлениеТовара"]
    metadata_columns = [h for h in headers if h not in EXCLUDED_COLUMNS and h not in text_columns]

    total_rows = ws.max_row - 1
    if max_records:
        total_rows = min(total_rows, max_records)

    result = []

    for i, row in enumerate(tqdm(ws.iter_rows(min_row=2), total=total_rows, desc="Конвертация")):
        if max_records and i >= max_records:
            break

        row_data = {
            header: str(cell.value).strip() if cell.value is not None else ""
            for header, cell in zip(headers, row)
        }

        # Основные данные для разметки
        data = {
            "text": row_data.get("ТоварПоставки", ""),
            "reference_text": row_data.get("ПредставлениеТовара", ""),
            "metadata": {}
        }

        # Добавляем все остальные колонки в metadata
        for column in metadata_columns:
            if column in row_data:
                data["metadata"][column] = row_data[column]

        result.append(data)

    with open(output_file, "w", encoding="utf-8") as f_out:
        json.dump(result, f_out, ensure_ascii=False, indent=2)

    wb.close()
    print(f"✅ Успешно сохранено {len(result)} записей в {output_file}")

# Пример использования:
convert_excel_to_datasaur(
    input_file="nlp_dataset.xlsx",
    output_file="datasaur_input.json",
    max_records=500  # None для всех записей
)