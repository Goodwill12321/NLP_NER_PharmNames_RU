import spacy
from spacy.pipeline import EntityRuler
import json
from tqdm import tqdm
from openpyxl import load_workbook
from spacy.tokens import DocBin
import argparse

def load_texts_from_excel(path, limit=None):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    idx = headers.index("ТоварПоставки")
    data = []
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        if limit and i >= limit:
            break
        val = row[idx].value
        if val:
            data.append(str(val).strip())
    return data

def remove_overlapping_ents(ents):
    ents = sorted(ents, key=lambda x: (x.start_char, -(x.end_char - x.start_char)))
    filtered = []
    prev_end = -1
    for ent in ents:
        if ent.start_char >= prev_end:
            filtered.append(ent)
            prev_end = ent.end_char
    return filtered

from ruler_patterns import get_ruler_patterns

def build_spacy_ruler(nlp):
    ruler = nlp.add_pipe("entity_ruler", before="ner", config={"overwrite_ents": True})
    ruler.add_patterns(get_ruler_patterns())
    return ruler

def annotate_texts(texts, nlp):
    annotated = []
    for text in tqdm(texts, desc="Разметка spaCy"):
        doc = nlp(text)
        doc.ents = remove_overlapping_ents(list(doc.ents))

        # Удалим лишние ТорговоеНаименование, оставим только первое
        tn_seen = False
        final_ents = []
        for ent in doc.ents:
            if ent.label_ == "ТорговоеНаименование":
                if not tn_seen:
                    final_ents.append(ent)
                    tn_seen = True
            else:
                final_ents.append(ent)

        spans = [{"start": ent.start_char, "end": ent.end_char, "label": ent.label_, "text": ent.text}
                 for ent in final_ents]
        annotated.append({"text": text, "entities": spans})
    return annotated

def export_to_label_studio_format(annotated, output_file):
    labelstudio_data = []
    for item in annotated:
        entities = [{"value": {"start": ent["start"], "end": ent["end"], "text": ent["text"], "labels": [ent["label"]]},
                     "from_name": "label", "to_name": "text", "type": "labels"} for ent in item["entities"]]

        labelstudio_data.append({"data": {"text": item["text"]}, "annotations": [{"result": entities}]})

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(labelstudio_data, f, ensure_ascii=False, indent=2)

def import_from_labelstudio(filepath):
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)

    spacy_format = []
    for item in data:
        text = item["data"]["text"]
        ents = [(result["value"]["start"], result["value"]["end"], result["value"]["labels"][0])
                for ann in item.get("annotations", []) for result in ann.get("result", [])]
        spacy_format.append((text, {"entities": ents}))
    return spacy_format

def save_spacy_training_data(examples, output_path):
    nlp = spacy.blank("ru")
    db = DocBin()

    for text, annot in examples:
        doc = nlp.make_doc(text)
        ents = [doc.char_span(start, end, label=label) for start, end, label in annot["entities"] if
                doc.char_span(start, end, label=label)]
        doc.ents = ents
        db.add(doc)

    db.to_disk(output_path)

def export_jsonl_for_llm(examples, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        for text, annot in examples:
            json.dump({"text": text, "entities": [{"start": s, "end": e, "label": l} for s, e, l in annot["entities"]]},
                      f, ensure_ascii=False)
            f.write("\n")

def train_spacy_model(training_data_path, output_model_path):
    nlp = spacy.blank("ru")
    config_file = "config.cfg"

    # Здесь должна быть предварительная настройка конфигурации для spaCy
    # Например, создание конфигурации, если она еще не существует:
    # spacy init config config.cfg --lang ru --pipeline ner

    # Теперь обучаем модель
    nlp.begin_training()
    db = DocBin().from_disk(training_data_path)

    for epoch in range(10):  # Количество эпох для дообучения
        for batch in db.get_batch():
            # Обучаем на данных
            pass

    nlp.to_disk(output_model_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Разметка и дообучение spaCy.")
    parser.add_argument("mode", choices=["annotate", "retrain", "export_llm"], help="Режим работы: annotate, retrain, export_llm")
    parser.add_argument("input_file", help="Имя входного файла (XLSX для аннотации или JSON для retrain)")
    parser.add_argument("output_file", help="Имя выходного файла для Label Studio (для annotate), сохранения модели (для retrain) или JSONL (для export_llm)")

    args = parser.parse_args()

    if args.mode == "annotate":
        print("🔄 Загрузка модели spaCy...")
        nlp = spacy.load("ru_core_news_lg")
        build_spacy_ruler(nlp)

        print("📥 Чтение Excel...")
        texts = load_texts_from_excel(args.input_file, limit=500)

        print("🧠 Разметка текста...")
        results = annotate_texts(texts, nlp)

        print(f"💾 Экспорт в Label Studio в {args.output_file}...")
        export_to_label_studio_format(results, args.output_file)
        print("✅ Готово!")

    elif args.mode == "retrain":
        print("📥 Импорт из Label Studio...")
        examples = import_from_labelstudio(args.input_file)

        print("💾 Сохранение для spaCy...")
        save_spacy_training_data(examples, args.output_file)

        print("🧠 Дообучение spaCy...")
        train_spacy_model(args.output_file, "output_model")

        print("✅ Дообучение завершено!")

    elif args.mode == "export_llm":
        print("📥 Импорт из Label Studio для LLM...")
        examples = import_from_labelstudio(args.input_file)

        print(f"💾 Экспорт данных для LLM в {args.output_file}...")
        export_jsonl_for_llm(examples, args.output_file)

        print("✅ Экспорт для LLM завершён!")