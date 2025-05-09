from openpyxl import load_workbook
import json
from tqdm import tqdm

def convert_to_labelstudio_json(input_file, output_file, max_records=None):
    EXCLUDED_COLUMNS = {
        "ТоварПоставки", "ПредставлениеТовара", "МНН", "МННСтрокой",
        "ЖН", "СМНН", "КЛП", "Наркотический", "ОКПД2",
        "РегНомерПредЦены", "ДатаОкончанияПредЦены",
        "ВидЗаписиРеестраЖН", "НомерРешенияРеестраЖН",
        "ДатаРегистрацииЦеныРеестраЖН", "ДатаВступленияВСилуРеестраЖН",
        "Дозировка_ЕИ_ПотребительскаяОКЕИ", "Штрихкод", 
        "ВладелецНаименование", "ВладелецСтрана", "ЛекформаДозировкаУпаковкаИзРеестраЖН", 
        "ВладелецРУИзРеестраЖН", "НомерРешенияРеестраЖН", "НомерРУ", "ДатаРУ"
    }

    print("🔍 Чтение Excel файла...")
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    total_rows = min(ws.max_row - 1, max_records) if max_records else ws.max_row - 1
    
    tasks = []
    
    for i, row in enumerate(tqdm(ws.iter_rows(min_row=2), total=total_rows, desc="Конвертация")):
        if max_records and i >= max_records:
            break
            
        row_data = {header: str(cell.value) if cell.value is not None else "" 
                   for header, cell in zip(headers, row)}
        
        tasks.append({
            "data": {
                "text": row_data.get("ТоварПоставки", ""),
                "meta": {
                    "reference": row_data.get("ПредставлениеТовара", ""),
                    **{k: v for k, v in row_data.items() 
                       if k not in EXCLUDED_COLUMNS}
                }
            }
        })
    
    with open(output_file, "w", encoding="utf-8") as f_out:
        json.dump({"tasks": tasks}, f_out, ensure_ascii=False, indent=2)
    
    wb.close()
    print(f"✅ Сохранено {len(tasks)} записей в {output_file}")

# Пример использования:
MAX_RECORDS = 100  # Задайте нужное количество записей или None для всех
convert_to_labelstudio_json(
    input_file="nlp_dataset.xlsx",
    output_file="output.json",
    max_records=MAX_RECORDS
)