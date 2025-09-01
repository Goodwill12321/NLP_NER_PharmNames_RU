import re
import json
from tqdm import tqdm
from openpyxl import load_workbook


def extract_heuristics(text):
    """Разбирает строку 'ТоварПоставки' и извлекает сущности."""
    result = {}

    # Торговое наименование (всё до первой дозировки или формы)
    name_match = re.match(r"(.+?)\s+(табл|капс|р-р|сусп|мазь|крем|супп|гель|аэроз|спрей|паст|лиоф|порош|рект|драже|инфуз|д/ин|д/рассасывания|гранул|амп|фл|тюб|саше|шпр|микрогранул)", text, re.IGNORECASE)
    if name_match:
        result["ТорговоеНаименование"] = name_match.group(1).strip()

    # Дозировка
    dose_match = re.search(r'(\d+[.,]?\d*)\s?(мг|мкг|г|мл|%)\b', text, re.IGNORECASE)
    if dose_match:
        result["Дозировка"] = f"{dose_match.group(1)} {dose_match.group(2)}".replace(',', '.').strip()
        result["ДозировкаКоличество"] = dose_match.group(1).replace(',', '.')
        result["Дозировка_ЕИ_Потребительская"] = dose_match.group(2)

    # ЛекФорма
    form_match = re.search(r'(табл|капс|р-р|сусп|мазь|крем|супп|гель|аэроз|спрей|паст|лиоф|порош|рект|драже|инфуз|д/ин|д/рассасывания|гранул|шпр|саше|микрогранул)', text, re.IGNORECASE)
    if form_match:
        result["ЛекФорма"] = form_match.group(0)

    # Первичная упаковка
    prim_match = re.search(r'(\d+)\s*(мл|мг|амп|фл|таб|капс|шт|тюб)', text, re.IGNORECASE)
    if prim_match:
        result["ПервичнаяУпаковкаКоличество"] = prim_match.group(1)
        result["ПервичнаяУпаковкаНазвание"] = prim_match.group(2)

    # Потребительская упаковка
    pot_match = re.search(r'№?\s*(\d+)', text)
    if pot_match:
        result["ПотребительскаяУпаковкаКолво"] = pot_match.group(1)

    return result


def parse_excel_with_heuristics(input_file, output_file, max_records=None):
    print("🔍 Чтение Excel файла...")
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    total_rows = ws.max_row - 1
    if max_records:
        total_rows = min(total_rows, max_records)

    results = []

    for i, row in enumerate(tqdm(ws.iter_rows(min_row=2), total=total_rows, desc="Обработка")):
        if max_records and i >= max_records:
            break

        row_data = {header: str(cell.value).strip() if cell.value is not None else "" for header, cell in zip(headers, row)}
        source_text = row_data.get("ТоварПоставки", "")
        if not source_text:
            continue

        annotations = extract_heuristics(source_text)
        results.append({
            "text": source_text,
            "entities": annotations
        })

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"✅ Сохранено {len(results)} записей в {output_file}")
    wb.close()


# Запуск
if __name__ == "__main__":
    parse_excel_with_heuristics(
        input_file="nlp_dataset.xlsx",
        output_file="parsed_heuristic_output.json",
        max_records=100  # Поменяй на None, чтобы обрабатывать все строки
    )
