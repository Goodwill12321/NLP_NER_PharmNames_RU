#dosn't use now
#conversion to label_studio format for NER task


from openpyxl import load_workbook
import json
from tqdm import tqdm

def convert_excel_to_labelstudio(input_file, output_file, max_records=None):
    EXCLUDED_COLUMNS = {
        "№", "ТоварПоставки", "ПредставлениеТовара", "МНН", "МННСтрокой", "ДатаРегистрацииПредЦены", "Дозировка_ЕИ_Потребительская_КодОКЕИ",
        "ЖН", "СМНН", "КЛП", "Наркотический", "ОКПД2", "Дозировка_ЕИ_КодОКЕИ",
        "РегНомерПредЦены", "ДатаОкончанияПредЦены",
        "ВидЗаписиРеестраЖН", "НомерРешенияРеестраЖН",
        "ДатаРегистрацииЦеныРеестраЖН", "ДатаВступленияВСилуРеестраЖН",
        "Дозировка_ЕИ_ПотребительскаяОКЕИ", "Штрихкод", 
        "ВладелецНаименование", "ВладелецСтрана", "ЛекформаДозировкаУпаковкаИзРеестраЖН", 
        "ВладелецРУИзРеестраЖН", "НомерРУ", "ДатаРУ"
    }

    print("🔍 Чтение Excel файла...")
    wb = load_workbook(filename=input_file, read_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    total_rows = ws.max_row - 1
    if max_records:
        total_rows = min(total_rows, max_records)

    result = []

    for i, row in enumerate(tqdm(ws.iter_rows(min_row=2), total=total_rows, desc="Конвертация")):
        if max_records and i >= max_records:
            break

        row_data = {
            header: str(cell.value).strip() if cell.value is not None else ""
            for header, cell in zip(headers, row)
        }

        # Собираем все нужные поля в один словарь data
        data = {
            "text": row_data.get("ТоварПоставки", ""),
            "reference": row_data.get("ПредставлениеТовара", "")
        }

        # Добавляем все оставшиеся поля (не исключенные и не text/reference)
        for k, v in row_data.items():
            if k not in EXCLUDED_COLUMNS and k not in ["ТоварПоставки", "ПредставлениеТовара", "№"]:
                data[k] = v

        result.append({"data": data})

    with open(output_file, "w", encoding="utf-8") as f_out:
        json.dump(result, f_out, ensure_ascii=False, indent=2)

    wb.close()
    print(f"✅ Успешно сохранено {len(result)} записей в {output_file}")

# Пример использования:
MAX_RECORDS = 100  # None для всех записей
convert_excel_to_labelstudio(
    input_file="nlp_dataset.xlsx",
    output_file="output.json",
    max_records=MAX_RECORDS
)
